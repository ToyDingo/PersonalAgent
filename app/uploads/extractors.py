from __future__ import annotations

import base64
from datetime import date, datetime
from pathlib import Path

from app.uploads.types import ExtractedContent, IcsEvent


def _extract_text_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _extract_text_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages: list[str] = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\n".join(pages)


def _extract_text_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    lines = [paragraph.text for paragraph in doc.paragraphs]
    return "\n".join(lines)


def _extract_text_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        lines.append(f"[Sheet: {sheet_name}]")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if values:
                lines.append(" | ".join(values))
    return "\n".join(lines)


def _to_iso_or_none(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def _extract_ics_events(path: Path) -> list[IcsEvent]:
    from icalendar import Calendar

    calendar = Calendar.from_ical(path.read_bytes())
    events: list[IcsEvent] = []
    for component in calendar.walk():
        if component.name != "VEVENT":
            continue
        dtstart_raw = component.get("dtstart")
        dtend_raw = component.get("dtend")
        start_val = getattr(dtstart_raw, "dt", None)
        end_val = getattr(dtend_raw, "dt", None)
        timezone = getattr(getattr(dtstart_raw, "params", {}), "get", lambda *_: None)("TZID")
        is_all_day = isinstance(start_val, date) and not isinstance(start_val, datetime)
        events.append(
            {
                "summary": str(component.get("summary")) if component.get("summary") else None,
                "dtstart": _to_iso_or_none(start_val),
                "dtend": _to_iso_or_none(end_val),
                "description": str(component.get("description")) if component.get("description") else None,
                "location": str(component.get("location")) if component.get("location") else None,
                "timezone": str(timezone) if timezone else None,
                "is_all_day": is_all_day,
            }
        )
    return events


def _extract_image(path: Path, extension: str) -> ExtractedContent:
    ext = extension.lower().strip()
    mime_type = "image/jpeg" if ext in {".jpg", ".jpeg"} else "image/png"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return {
        "type": "image",
        "content_base64": encoded,
        "mime_type": mime_type,
    }


def extract_content_from_file(path_str: str, extension: str) -> ExtractedContent:
    path = Path(path_str)
    if not path.exists():
        raise FileNotFoundError(path_str)
    ext = extension.lower().strip()
    if ext == ".txt":
        return {"type": "text", "content": _extract_text_txt(path)}
    if ext == ".pdf":
        return {"type": "text", "content": _extract_text_pdf(path)}
    if ext == ".docx":
        return {"type": "text", "content": _extract_text_docx(path)}
    if ext == ".xlsx":
        return {"type": "text", "content": _extract_text_xlsx(path)}
    if ext == ".ics":
        return {"type": "ics_events", "events": _extract_ics_events(path)}
    if ext in {".png", ".jpg", ".jpeg"}:
        return _extract_image(path, ext)
    raise ValueError(f"Unsupported extractor extension: {extension}")

